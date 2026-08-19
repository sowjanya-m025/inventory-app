"""
Exports the full inventory dataset from Postgres into a single Excel
workbook, one sheet per table, each formatted as a proper Excel Table
(not just a raw range) — this is what lets Power BI Service treat each
sheet as a clean, structured dataset once uploaded.

Run from the inventory-api folder (venv active):
    python3 ml/export_to_excel.py

Output: inventory_export.xlsx (in the current folder)
"""
import os
import sys

import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import engine

OUTPUT_FILE = "inventory_export.xlsx"

# Each query pulls a clean, join-friendly table — Power BI will build
# relationships between these using product_id / warehouse_id.
QUERIES = {
    "Products": """
        SELECT product_id, sku, name, category_id, supplier_id,
               unit_price, unit_cost, reorder_point, reorder_qty,
               is_active, created_at
        FROM products
        ORDER BY product_id
    """,
    "Categories": "SELECT category_id, name FROM categories ORDER BY category_id",
    "Suppliers": "SELECT supplier_id, name, contact_email FROM suppliers ORDER BY supplier_id",
    "Warehouses": "SELECT warehouse_id, name, location FROM warehouses ORDER BY warehouse_id",
    "StockTransactions": """
        SELECT transaction_id, product_id, warehouse_id, transaction_type,
               quantity, transaction_date
        FROM stock_transactions
        ORDER BY transaction_date
    """,
    "InventorySnapshots": """
        SELECT snapshot_id, product_id, warehouse_id, snapshot_date,
               quantity_on_hand, units_sold, units_received
        FROM inventory_snapshots
        ORDER BY snapshot_date
    """,
    "DemandForecasts": """
        SELECT forecast_id, product_id, warehouse_id, forecast_date,
               predicted_demand, lower_bound, upper_bound, model_version
        FROM demand_forecasts
        ORDER BY forecast_date
    """,
    "CurrentInventory": """
        SELECT inventory_id, product_id, warehouse_id, quantity_on_hand, updated_at
        FROM inventory
        ORDER BY product_id
    """,
}


def export():
    print("Connecting to database and pulling tables...")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, query in QUERIES.items():
            df = pd.read_sql(query, engine)

            # Excel sheet/table names can't exceed 31 chars or contain spaces in table names
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  {sheet_name}: {len(df)} rows")

    # Second pass: convert each sheet's data range into a real Excel Table
    # (ListObject) with a named style — this is what makes Power BI Service
    # recognize it as structured data rather than a flat range.
    from openpyxl import load_workbook

    wb = load_workbook(OUTPUT_FILE)
    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    for sheet_name in QUERIES.keys():
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue  # skip empty tables, Excel Tables need at least one data row

        max_col_letter = get_column_letter(ws.max_column)
        table_ref = f"A1:{max_col_letter}{ws.max_row}"

        table = Table(displayName=f"tbl{sheet_name}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(table)

        for row in ws.iter_rows():
            for cell in row:
                cell.font = header_font if cell.row == 1 else body_font

        # Auto-width columns for readability
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    wb.save(OUTPUT_FILE)
    print(f"\nDone. Wrote {OUTPUT_FILE} with {len(QUERIES)} formatted tables.")
    print("Upload this file to Power BI Service (app.powerbi.com) to build your dashboards.")


if __name__ == "__main__":
    export()
